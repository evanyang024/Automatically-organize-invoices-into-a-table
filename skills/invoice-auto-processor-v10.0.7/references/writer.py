"""Excel 写入模块 — 金额为数字、无边框、多明细合并 + 自动保存关闭已打开Excel"""
import os, re, json as _json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def get_output_path(output_dir, filename):
    return os.path.join(output_dir, filename)


def init_or_load_workbook(filepath, columns):
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "发票台账"
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    return wb, ws


def append_invoice_rows(filepath, columns, invoice_data):
    # 如果 Excel 已被打开，先用 AppleScript 保存并关闭
    _close_excel_if_open(os.path.basename(filepath))
    
    wb, ws = init_or_load_workbook(filepath, columns)
    items = invoice_data.get("items", [])
    next_row = _find_empty_row(ws)

    if items:
        # 一张发票只写一行
        item = items[0]
        if len(items) > 1:
            item["货物或服务名称"] = f"{item['货物或服务名称']}等{len(items)}项"
        row_data = _build_row(columns, item, invoice_data)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            if col_idx == 2 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
    else:
        row_data = _build_row(columns, {}, invoice_data)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    # 按项目名称分类排序，同类型发票挨在一起
    _sort_rows_by_name(ws)

    wb.save(filepath)
    wb.close()


def _sort_rows_by_name(ws):
    """按第一列（货物或服务名称）排序，同类型发票挨在一起"""
    rows_data = []
    for row in range(2, ws.max_row + 1):
        row_values = []
        has_data = False
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            row_values.append(val)
            if val is not None:
                has_data = True
        if not has_data:
            continue
        rows_data.append(row_values)
    
    if not rows_data:
        return
    
    rows_data.sort(key=lambda r: str(r[0]) if r[0] is not None else "zzz")
    
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    for i, row_data in enumerate(rows_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            if col_idx == 2 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'


def _find_empty_row(ws):
    """找到第一个空行（所有列都没有数据）"""
    # 从第2行开始找（第1行是表头）
    for row in range(2, ws.max_row + 2):
        is_empty = True
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value is not None:
                is_empty = False
                break
        if is_empty:
            return row
    return ws.max_row + 1


def _build_row(columns, item, invoice_data):
    import re
    mapping = {
        "货物或服务名称": item.get("货物或服务名称", ""),
        "规格型号": item.get("规格型号", ""),
        "数量": item.get("数量", ""),
        "单价": item.get("单价", ""),
        "金额": item.get("金额", ""),
        "税率": item.get("税率", ""),
        "不含税金额合计": invoice_data.get("amount_excluding_tax", ""),
        "税额合计": invoice_data.get("total_tax", ""),
        "价税合计": _to_number(invoice_data.get("total_amount", "")),
        "开票日期": re.sub(r'[0-9]{4}[ ]*年[ ]*', '', invoice_data.get("invoice_date", "")),
        "备注": invoice_data.get("remarks", ""),
    }
    return [mapping.get(col, "") for col in columns]


def _to_number(val):
    if val:
        try:
            return float(str(val).replace(",", ""))
        except ValueError:
            return val
    return val


def _close_excel_if_open(target_filename):
    """检测 Excel/WPS 是否打开了目标文件，如果有锁则尝试关闭"""
    import subprocess, platform
    if platform.system() != "Darwin":
        return
    
    # 先检测文件是否有锁：尝试用 openpyxl 读写
    try:
        from openpyxl import load_workbook
        wb = load_workbook(target_filename) if os.path.exists(target_filename) else None
        if wb:
            wb.close()
            return  # 文件没有被锁定
    except Exception:
        pass  # 文件被锁定，需要关闭
    
    # 文件被锁定，尝试通过 System Events 激活并关闭
    for app_name in ["Microsoft Excel", "wpsoffice"]:
        try:
            script = f'''
            tell application "System Events"
                tell process "{app_name}"
                    set frontmost to true
                    keystroke "s" using command down
                    delay 0.3
                    keystroke "w" using command down
                end tell
            end tell
            '''
            result = subprocess.run(["osascript", "-e", script], capture_output=True, timeout=10)
            if result.returncode == 0:
                break
        except Exception:
            continue
    
    # 等待文件解锁
    import time
    time.sleep(1)


# === 去重跟踪 ===
def _tracking_file(output_dir):
    return os.path.join(output_dir, ".processed_invoices.json")


def get_processed_invoice_numbers(output_dir):
    tf = _tracking_file(output_dir)
    if not os.path.exists(tf):
        return set()
    try:
        with open(tf, "r", encoding="utf-8") as f:
            return set(_json.load(f))
    except Exception:
        return set()


def mark_invoice_processed(output_dir, invoice_number):
    if not invoice_number:
        return
    tf = _tracking_file(output_dir)
    existing = get_processed_invoice_numbers(output_dir)
    existing.add(invoice_number)
    with open(tf, "w", encoding="utf-8") as f:
        _json.dump(sorted(existing), f, ensure_ascii=False)
